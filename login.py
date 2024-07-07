import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import cv2
from PIL import Image, ImageTk
import ctypes
from openpyxl import load_workbook
import os
import importlib

class LoginPage:
    def __init__(self, root):
        self.root = root
        self.root.title("Login Page")

        user32 = ctypes.windll.user32
        self.screen_width = user32.GetSystemMetrics(0)
        self.screen_height = user32.GetSystemMetrics(1)

        self.cap = cv2.VideoCapture(r"C:\Users\dushyanth m\Downloads\Stock Market Background Video(720P_HD).mp4")

        self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

        self.scale_factor = min(self.screen_width / self.video_width, self.screen_height / self.video_height)

        self.canvas = tk.Canvas(root, width=self.video_width * self.scale_factor, height=self.video_height * self.scale_factor)
        self.canvas.pack()

        style = ttk.Style()
        style.configure('Custom.TLabel', font=('Helvetica', 14, 'bold'), foreground='blue')
        style.configure('Custom.TEntry', font=('Helvetica', 14, 'bold'))
        style.configure('Custom.TButton', font=('Helvetica', 14, 'bold'), foreground='white', background='#336699')
        style.configure('Login.TButton', font=('Helvetica', 14, 'bold'), foreground='black', background='red')
        style.configure('Signup.TButton', font=('Helvetica', 14, 'bold'), foreground='black', background='green')

        self.username_label = ttk.Label(root, text="Username:", style='Custom.TLabel')
        self.username_label.place(relx=0.7, rely=0.4, anchor="center")
        self.username_entry = ttk.Entry(root, style='Custom.TEntry')
        self.username_entry.place(relx=0.8, rely=0.4, anchor="center")
        self.password_label = ttk.Label(root, text="Password:", style='Custom.TLabel')
        self.password_label.place(relx=0.7, rely=0.5, anchor="center")
        self.password_entry = ttk.Entry(root, show="*", style='Custom.TEntry')
        self.password_entry.place(relx=0.8, rely=0.5, anchor="center")
        self.phone_label = ttk.Label(root, text="Phone Number:", style='Custom.TLabel')
        self.phone_label.place(relx=0.7, rely=0.6, anchor="center")
        self.phone_entry = ttk.Entry(root, style='Custom.TEntry')
        self.phone_entry.place(relx=0.8, rely=0.6, anchor="center")
        self.login_button = ttk.Button(root, text="Login", command=self.handle_login, style='Login.TButton')
        self.login_button.place(relx=0.75, rely=0.8, anchor="center")
        self.signup_button = ttk.Button(root, text="Sign Up", command=self.handle_signup, style='Signup.TButton')
        self.signup_button.place(relx=0.85, rely=0.8, anchor="center")

        self.play_video()

        self.root.geometry(f"{self.screen_width}x{self.screen_height}")

        self.credentials_filename = r"C:\Users\dushyanth m\Documents\PREDICTIVE ANALYSI\credentials.xlsx"
        self.credentials = self.load_credentials()

    def play_video(self):
        ret, frame = self.cap.read()
        if ret:
            frame = cv2.resize(frame, (int(self.video_width * self.scale_factor), int(self.video_height * self.scale_factor)))
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame)
            imgtk = ImageTk.PhotoImage(image=img)
            self.canvas.imgtk = imgtk
            self.canvas.create_image(0, 0, anchor=tk.NW, image=imgtk)
            self.root.after(33, self.play_video)
        else:
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.play_video()

    def load_credentials(self):
        credentials = {}
        if os.path.exists(self.credentials_filename):
            wb = load_workbook(self.credentials_filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                username, password, phonenumber = row
                credentials[username] = (password, phonenumber)
        return credentials

    def save_credentials(self, username, password, phonenumber):
        wb = load_workbook(self.credentials_filename)
        ws = wb.active
        row = (username, password, phonenumber)
        ws.append(row)
        wb.save(self.credentials_filename)

    def login(self, username, password):
        credentials = self.load_credentials()
        if username in credentials and credentials[username][0] == password:
            return True, credentials[username][1]
        else:
            return False, None

    def signup(self, username, password, phonenumber):
        credentials = self.load_credentials()
        if username not in credentials:
            credentials[username] = (password, phonenumber)
            self.save_credentials(username, password, phonenumber)
            return True
        else:
            return False

    def handle_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        success, phone_number = self.login(username, password)
        if success:
            self.root.withdraw()
            self.open_panel(username)
            self.root.destroy()
        else:
            messagebox.showerror("Login", "Invalid username or password.")

    def handle_signup(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        phonenumber = self.phone_entry.get()

        if not phonenumber.isdigit() or len(phonenumber) != 10:
            messagebox.showerror("Sign Up", "Please enter a valid 10-digit phone number.")
            return

        if self.signup(username, password, phonenumber):
            messagebox.showinfo("Sign Up", "Sign up successful!")
        else:
            messagebox.showerror("Sign Up", "Username already exists.")

    def open_panel(self, username):
        panel = UserPanel(self.root, username, self)
        panel.panel.mainloop()

class UserPanel:
    def import_Alcou(self):
        try:
            import Alcou
            print("Alcou package imported successfully.")
        except ImportError as e:
            messagebox.showerror("Import Error", f"Failed to import Alcou: {e}")
            
    def import_TFtest(self):
        try:
            import TFtest
            print("TFtest package imported successfully.")
        except ImportError as e:
            messagebox.showerror("Import Error", f"Failed to import TFtest: {e}")
            
    def import_clustering(self):
        try:
            import clustering
            print("clustering package imported successfully.")
        except ImportError as e:
            messagebox.showerror("Import Error", f"Failed to import clustering: {e}")

    def __init__(self, root, username, login_page):
        self.username = username
        self.root = root
        self.login_page = login_page

        self.panel = tk.Toplevel(root)
        self.panel.title("CURRENCY SWAP ")

        self.screen_width = self.panel.winfo_screenwidth()
        self.screen_height = self.panel.winfo_screenheight()
        self.panel.geometry(f"{self.screen_width}x{self.screen_height}")

        bg_img = Image.open(r"C:\Users\dushyanth m\Documents\PREDICTIVE ANALYSI\bg-1.jpg")
        bg_img = bg_img.resize((self.screen_width, self.screen_height), Image.LANCZOS)
        self.bg_img_tk = ImageTk.PhotoImage(bg_img)

        self.bg_label = tk.Label(self.panel, image=self.bg_img_tk)
        self.bg_label.place(relx=0, rely=0, relwidth=1, relheight=1)

        style = ttk.Style()
        style.configure('Custom.TLabel', font=('Helvetica', 14, 'bold'), foreground='black')
        style.configure('Custom.TEntry', font=('Helvetica', 14, 'bold'))
        style.configure('Custom.TButton', font=('Helvetica', 14, 'bold'), foreground='black', background='blue')

        self.forex_btn = tk.Button(self.panel, text="Forex", command=self.show_forex_options, font=('Helvetica', 14, 'bold'))
        self.forex_btn.place(relx=0.1, rely=0.1, anchor=tk.CENTER)

        self.pca_btn = tk.Button(self.panel, text="PCA", command=self.show_pca_options, font=('Helvetica', 14, 'bold'))
        self.pca_btn.place(relx=0.2, rely=0.1, anchor=tk.CENTER)

        self.pca_btnO = tk.Button(self.panel, text="OTHER CURRENCY", command=self.import_Alcou, font=('Helvetica', 14, 'bold'))
        self.pca_btnO.place(relx=0.35, rely=0.1, anchor=tk.CENTER)
        
        self.regression1_btn = tk.Button(self.panel, text=" Regression ", command=self.perform_regression1, font=('Helvetica', 14, 'bold'))
        self.regression1_btn.place(relx=0.5, rely=0.1, anchor=tk.CENTER)

        self.test_btn = tk.Button(self.panel, text=" TEST ", command=self.import_TFtest, font=('Helvetica', 14, 'bold'))
        self.test_btn.place(relx=0.65, rely=0.1, anchor=tk.CENTER)

        self.clustering_btn = tk.Button(self.panel, text="Clustering", command=self.import_clustering, font=('Helvetica', 14, 'bold'))
        self.clustering_btn.place(relx=0.65, rely=0.4, anchor=tk.CENTER)

        self.factor_btn = tk.Button(self.panel, text="Factor Analysis", command=self.show_fa_options, font=('Helvetica', 14, 'bold'))
        self.factor_btn.place(relx=0.2, rely=0.4, anchor=tk.CENTER)
        

        self.logout_btn = tk.Button(self.panel, text="Logout", command=self.logout, font=('Helvetica', 14, 'bold'))
        self.logout_btn.place(relx=0.5, rely=0.8, anchor=tk.CENTER)

        self.proceed_btn = tk.Button(self.panel, text="Proceed", command=self.proceed_function, font=('Helvetica', 14, 'bold'))
        self.proceed_btn.place(relx=0.1, rely=0.3, anchor=tk.CENTER)  
        self.proceed_btn.place_forget()
        
        self.fa_btn = tk.Button(self.panel, text="FA", command=self.fa_function, font=('Helvetica', 14, 'bold'))
        self.fa_btn.place(relx=0.6, rely=0.6, anchor=tk.CENTER)
        self.fa_btn.place_forget()

        self.result_btn = tk.Button(self.panel, text="Result", command=self.result_function, font=('Helvetica', 14, 'bold'))
        self.result_btn.place(relx=0.2, rely=0.3, anchor=tk.CENTER)
        self.result_btn.place_forget()

        messagebox.showinfo("Welcome", f"Welcome, {username}!")



        
    def show_forex_options(self):
        options = ["Europe", "Australia", "Switzerland", "Japan", "USD"]

        self.selected_forex_option = tk.StringVar(self.panel)
        self.selected_forex_option.set(options[0])

        forex_dropdown = tk.OptionMenu(self.panel, self.selected_forex_option, *options)
        forex_dropdown.place(relx=0.1, rely=0.2, anchor=tk.CENTER)

        self.proceed_btn.place(relx=0.1, rely=0.3, anchor=tk.CENTER)
        

    def show_pca_options(self):
        options = ["Europe", "Australia", "Switzerland", "Japan", "USD"]

        self.selected_pca_option = tk.StringVar(self.panel)
        self.selected_pca_option.set(options[0])

        pca_dropdown = tk.OptionMenu(self.panel, self.selected_pca_option, *options)
        pca_dropdown.place(relx=0.2, rely=0.2, anchor=tk.CENTER)

        self.result_btn.place(relx=0.2, rely=0.3, anchor=tk.CENTER)

    def show_fa_options(self):
        options = ["Europe", "Australia", "Switzerland", "Japan", "USD","OVERALL"]

        self.selected_fa_option = tk.StringVar(self.panel)
        self.selected_fa_option.set(options[0])

        fa_dropdown  = tk.OptionMenu(self.panel, self.selected_fa_option, *options)
        fa_dropdown.place(relx=0.2, rely=0.6, anchor=tk.CENTER)

        self.fa_btn.place(relx=0.2, rely=0.5, anchor=tk.CENTER)
        
    def perform_regression1(self):
       self.regression_type = tk.StringVar(self.panel)
       self.regression_type.set("Polynomial Regression")

       options = ["Polynomial Regression", "Multilinear Regression"]

       regression_dropdown = tk.OptionMenu(self.panel, self.regression_type, *options)
       regression_dropdown.place(relx=0.5, rely=0.2, anchor=tk.CENTER)
       
       self.analysis_btn = tk.Button(self.panel, text="Analyze", command=self.perform_analysis, font=('Helvetica', 14, 'bold'))
       self.analysis_btn.place(relx=0.5, rely=0.3, anchor=tk.CENTER)
        

    def proceed_function(self):
        selected_forex_option = self.selected_forex_option.get()
        print(f"Proceeding with Forex option: {selected_forex_option}")
        if selected_forex_option == "Europe":
            importlib.import_module("criceur")
        elif selected_forex_option == "Australia":
            importlib.import_module("cricaus")
        elif selected_forex_option == "Switzerland":
            importlib.import_module("cricswi")
        elif selected_forex_option == "Japan":
            importlib.import_module("cricjap")
        elif selected_forex_option == "USD":
            importlib.import_module("cricusd")
        else:
            return selected_forex_option
        
    def fa_function(self):
       selected_fa_option = self.selected_fa_option.get()
       print(f"Proceeding with fa option: {selected_fa_option}")
       if selected_fa_option == "Europe":
            importlib.import_module("fa-eur")
       elif selected_fa_option == "Australia":
            importlib.import_module("fa-aud")
       elif selected_fa_option == "Switzerland":
            importlib.import_module("fa-chf")
       elif selected_fa_option == "Japan":
            importlib.import_module("fa-jpy")
       elif selected_fa_option == "USD":
            importlib.import_module("fa-usd")
       elif selected_fa_option == "OVERALL":
            importlib.import_module("factorfor5countries")
       else:
            return selected_fa_option
     
        
   

    def result_function(self):
        selected_pca_option = self.selected_pca_option.get()
        print(f"Showing result for PCA option: {selected_pca_option}")
        if selected_pca_option == "Europe":
            importlib.import_module("Eurcrisis")
        elif selected_pca_option == "Australia":
            importlib.import_module("Auscrisis")
        elif selected_pca_option == "Switzerland":
            importlib.import_module("swicrisis")
        elif selected_pca_option == "Japan":
            importlib.import_module("Japcrisis")
        elif selected_pca_option == "USD":
            importlib.import_module("Amecrisis")
        else:
            return selected_pca_option
        
    def perform_analysis(self):
        selected_regression_type = self.regression_type.get()
        print(f"Performing analysis for regression type: {selected_regression_type}")
        if selected_regression_type =="Polynomial Regression":
            importlib.import_module("polynomialregression")
        elif selected_regression_type =="Multilinear Regression":
            importlib.import_module("multilinearregression")
            
            

    def logout(self):
        self.panel.destroy()
        self.root.deiconify()
        self.login_page.root.deiconify()
        self.login_page.username_entry.delete(0, tk.END)
        self.login_page.password_entry.delete(0, tk.END)

def main():
    root = tk.Tk()
    login_page = LoginPage(root)
    root.mainloop()

if __name__ == "__main__":
    main()
